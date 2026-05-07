from openpyxl import load_workbook
import json

# Load the Excel file
wb = load_workbook('NQ v2.8 (2).xlsx')
ws = wb.active

# Extract headers
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(1, col).value
    if cell_value:
        headers.append(str(cell_value))

print("=" * 100)
print(f"Sheet Name: {ws.title}")
print(f"Total Rows: {ws.max_row} | Total Columns: {ws.max_column}")
print("=" * 100)
print("\nCOLUMN HEADERS:")
for i, header in enumerate(headers, 1):
    print(f"{i}. {header}")

print("\n" + "=" * 100)
print("FIRST 15 ROWS OF DATA:")
print("=" * 100 + "\n")

# Print first 15 rows with all columns
for row_idx in range(2, min(17, ws.max_row + 1)):
    for col_idx in range(1, min(20, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        value = cell.value
        if value is not None:
            # Try to detect if it's a formula
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                print(f"Row {row_idx}, Col {col_idx}: FORMULA = {cell.value}")
            else:
                print(f"{headers[col_idx-1] if col_idx <= len(headers) else f'Col{col_idx}'}: {value}", end=" | ")
    print()

print("\n" + "=" * 100)
print("ANALYSIS - Pricing Related Columns:")
print("=" * 100 + "\n")

# Look for pricing related columns
pricing_cols = {}
for col_idx, header in enumerate(headers, 1):
    header_lower = header.lower()
    if any(keyword in header_lower for keyword in ['price', 'cost', 'rate', 'formula', 'charge', 'fee', 'margin', 'markup']):
        pricing_cols[col_idx] = header
        print(f"Column {col_idx}: {header}")
        # Print values from this column
        print("  Values (first 5):")
        for row_idx in range(2, min(7, ws.max_row + 1)):
            cell = ws.cell(row_idx, col_idx)
            print(f"    Row {row_idx}: {cell.value}")
        print()

wb.close()
